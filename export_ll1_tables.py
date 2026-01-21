from __future__ import annotations

"""
Export LL(1) lab artifacts (grammar, FIRST, FOLLOW, parse table) into Excel-friendly files.

Outputs (always):
  - LL1_Grammar.csv
  - LL1_FIRST.csv
  - LL1_FOLLOW.csv
  - LL1_ParseTable.csv

Optional (only if openpyxl is installed):
  - LL1_Parse_Table.xlsx  (multiple sheets)

Run:
  python -X utf8 export_ll1_tables.py
"""

import csv
from pathlib import Path
from typing import Dict, List, Set

from webapp.ll1 import (
	EOF,
	EPS,
	Grammar,
	Production,
	build_ll1_table,
	compute_first_sets,
	compute_follow_sets,
	default_assignment_expr_grammar,
)


ROOT = Path(__file__).resolve().parent


def fmt_sym(s: str) -> str:
	return "eps" if s == EPS else s


def fmt_prod(p: Production) -> str:
	return str(p).replace(EPS, "eps")


def export_grammar_csv(grammar: Grammar) -> None:
	out_path = ROOT / "LL1_Grammar.csv"
	with out_path.open("w", newline="", encoding="utf-8") as f:
		w = csv.writer(f)
		w.writerow(["Section", "Production"])
		w.writerow(["Original (left-recursive) common form", "E -> E + T | T"])
		w.writerow(["Original (left-recursive) common form", "T -> T * F | F"])
		w.writerow(["Original (left-recursive) common form", "F -> ( E ) | id | num"])
		w.writerow([])
		w.writerow(["LL(1) equivalent used in this project", "S -> id = E ;"])
		w.writerow(["LL(1) equivalent used in this project", "E -> T E'"])
		w.writerow(["LL(1) equivalent used in this project", "E' -> + T E' | eps"])
		w.writerow(["LL(1) equivalent used in this project", "T -> F T'"])
		w.writerow(["LL(1) equivalent used in this project", "T' -> * F T' | eps"])
		w.writerow(["LL(1) equivalent used in this project", "F -> ( E ) | id | num"])
		w.writerow([])
		w.writerow(["Productions (expanded, one per line)", ""])
		for p in grammar.productions:
			w.writerow(["", fmt_prod(p)])


def export_set_csv(filename: str, title: str, sets: Dict[str, Set[str]]) -> None:
	out_path = ROOT / filename
	with out_path.open("w", newline="", encoding="utf-8") as f:
		w = csv.writer(f)
		w.writerow([title, "Symbols (sorted)"])
		for nt, syms in sorted(sets.items()):
			w.writerow([nt, " ".join(sorted(fmt_sym(s) for s in syms))])


def export_parse_table_csv(grammar: Grammar, table: Dict[str, Dict[str, Production]]) -> None:
	terms = sorted(list(grammar.terminals | {EOF}))
	nts = sorted(list(grammar.nonterminals))

	out_path = ROOT / "LL1_ParseTable.csv"
	with out_path.open("w", newline="", encoding="utf-8") as f:
		w = csv.writer(f)
		w.writerow(["NonTerminal"] + terms)
		for nt in nts:
			row: List[str] = [nt]
			for t in terms:
				p = table.get(nt, {}).get(t)
				row.append(fmt_prod(p) if p is not None else "")
			w.writerow(row)


def try_export_xlsx(grammar: Grammar, first: Dict[str, Set[str]], follow: Dict[str, Set[str]], table: Dict[str, Dict[str, Production]]) -> bool:
	try:
		import openpyxl  # type: ignore
		from openpyxl.utils import get_column_letter  # type: ignore
	except Exception:
		return False

	wb = openpyxl.Workbook()
	wb.remove(wb.active)

	# Grammar sheet
	ws = wb.create_sheet("Grammar")
	ws.append(["Section", "Production"])
	ws.append(["Original (left-recursive) common form", "E -> E + T | T"])
	ws.append(["Original (left-recursive) common form", "T -> T * F | F"])
	ws.append(["Original (left-recursive) common form", "F -> ( E ) | id | num"])
	ws.append([])
	ws.append(["LL(1) equivalent used", "S -> id = E ;"])
	ws.append(["LL(1) equivalent used", "E -> T E'"])
	ws.append(["LL(1) equivalent used", "E' -> + T E' | eps"])
	ws.append(["LL(1) equivalent used", "T -> F T'"])
	ws.append(["LL(1) equivalent used", "T' -> * F T' | eps"])
	ws.append(["LL(1) equivalent used", "F -> ( E ) | id | num"])
	ws.append([])
	ws.append(["Productions (expanded)", ""])
	for p in grammar.productions:
		ws.append(["", fmt_prod(p)])

	# FIRST sheet
	ws = wb.create_sheet("FIRST")
	ws.append(["NonTerminal", "Symbols (sorted)"])
	for nt, syms in sorted(first.items()):
		ws.append([nt, " ".join(sorted(fmt_sym(s) for s in syms))])

	# FOLLOW sheet
	ws = wb.create_sheet("FOLLOW")
	ws.append(["NonTerminal", "Symbols (sorted)"])
	for nt, syms in sorted(follow.items()):
		ws.append([nt, " ".join(sorted(fmt_sym(s) for s in syms))])

	# Parse table sheet
	ws = wb.create_sheet("ParseTable")
	terms = sorted(list(grammar.terminals | {EOF}))
	nts = sorted(list(grammar.nonterminals))
	ws.append(["NonTerminal"] + terms)
	for nt in nts:
		row = [nt]
		for t in terms:
			p = table.get(nt, {}).get(t)
			row.append(fmt_prod(p) if p is not None else "")
		ws.append(row)

	# Basic column sizing
	for sheet in wb.worksheets:
		for col in range(1, sheet.max_column + 1):
			letter = get_column_letter(col)
			sheet.column_dimensions[letter].width = 22 if col == 1 else 18

	out_path = ROOT / "LL1_Parse_Table.xlsx"
	wb.save(out_path)
	return True


def main() -> None:
	grammar = default_assignment_expr_grammar()
	first = compute_first_sets(grammar)
	follow = compute_follow_sets(grammar, first)
	table, conflicts = build_ll1_table(grammar, first, follow)

	# Export CSVs (Excel can open these directly)
	export_grammar_csv(grammar)
	export_set_csv("LL1_FIRST.csv", "FIRST", first)
	export_set_csv("LL1_FOLLOW.csv", "FOLLOW", follow)
	export_parse_table_csv(grammar, table)

	# Optional XLSX
	xlsx_ok = try_export_xlsx(grammar, first, follow, table)

	print("Wrote:", "LL1_Grammar.csv, LL1_FIRST.csv, LL1_FOLLOW.csv, LL1_ParseTable.csv")
	print("LL(1) conflicts:", conflicts)
	print("Wrote LL1_Parse_Table.xlsx:", xlsx_ok)


if __name__ == "__main__":
	main()
